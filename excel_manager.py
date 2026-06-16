import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os

class ExcelCalendarManager:
    def __init__(self, filename="content_calendar.xlsx"):
        self.filename = filename
        if not os.path.exists(filename):
            self.create_new_file()
        else:
            self.workbook = openpyxl.load_workbook(filename)
            self.sheet = self.workbook.active
    
    def create_new_file(self):
        """Create a new Excel file with the required structure"""
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Content Calendar"
        
        # Define headers
        headers = ['DATE', 'MAIN PAGE', 'STATUS', 'CANADA LC COMMUNITY', 
                   'STATUS', 'YOUTUBE', 'STATUS']
        
        # Apply headers with styling
        for col, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Adjust column widths
        column_widths = [12, 25, 15, 25, 15, 25, 15]
        for i, width in enumerate(column_widths, 1):
            self.sheet.column_dimensions[chr(64 + i)].width = width
        
        self.workbook.save(self.filename)
    
    def add_entry(self, date, main_page, main_status, canada_content, 
                  canada_status, youtube_content, youtube_status):
        """Add a new entry to the calendar"""
        # Find the next empty row
        next_row = self.sheet.max_row + 1
        
        entries = [date, main_page, main_status, canada_content, 
                  canada_status, youtube_content, youtube_status]
        
        for col, value in enumerate(entries, 1):
            cell = self.sheet.cell(row=next_row, column=col, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Apply borders
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col in range(1, 8):
            self.sheet.cell(row=next_row, column=col).border = thin_border
        
        self.workbook.save(self.filename)
        return True
    
    def get_all_entries(self):
        """Get all entries from the sheet"""
        entries = []
        for row in range(2, self.sheet.max_row + 1):
            entry = {
                'row': row,
                'date': self.sheet.cell(row=row, column=1).value,
                'main_page': self.sheet.cell(row=row, column=2).value,
                'main_status': self.sheet.cell(row=row, column=3).value,
                'canada_content': self.sheet.cell(row=row, column=4).value,
                'canada_status': self.sheet.cell(row=row, column=5).value,
                'youtube_content': self.sheet.cell(row=row, column=6).value,
                'youtube_status': self.sheet.cell(row=row, column=7).value
            }
            entries.append(entry)
        return entries
    
    def update_entry(self, row, date, main_page, main_status, canada_content, 
                    canada_status, youtube_content, youtube_status):
        """Update an existing entry"""
        updates = [date, main_page, main_status, canada_content, 
                  canada_status, youtube_content, youtube_status]
        
        for col, value in enumerate(updates, 1):
            self.sheet.cell(row=row, column=col, value=value)
        
        self.workbook.save(self.filename)
        return True
    
    def delete_entry(self, row):
        """Delete an entry"""
        self.sheet.delete_rows(row)
        self.workbook.save(self.filename)
        return True
    
    def search_entries(self, keyword):
        """Search entries by keyword"""
        results = []
        entries = self.get_all_entries()
        
        for entry in entries:
            if any(keyword.lower() in str(value).lower() for value in entry.values() 
                   if value is not None):
                results.append(entry)
        
        return results
    
    def filter_by_status(self, status):
        """Filter entries by status"""
        results = []
        entries = self.get_all_entries()
        
        for entry in entries:
            if (entry['main_status'] == status or 
                entry['canada_status'] == status or 
                entry['youtube_status'] == status):
                results.append(entry)
        
        return results