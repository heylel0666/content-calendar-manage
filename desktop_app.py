import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from datetime import datetime

class ContentCalendarApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Content Calendar Manager")
        self.root.geometry("1200x600")
        
        # File handling
        self.filename = "content_calendar.xlsx"
        self.load_or_create_file()
        
        # Create UI
        self.create_widgets()
        self.load_data()
    
    def load_or_create_file(self):
        if not os.path.exists(self.filename):
            self.create_new_file()
        else:
            self.workbook = openpyxl.load_workbook(self.filename)
            self.sheet = self.workbook.active
    
    def create_new_file(self):
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        headers = ['DATE', 'MAIN PAGE', 'STATUS', 'CANADA LC COMMUNITY', 
                   'STATUS', 'YOUTUBE', 'STATUS']
        
        for col, header in enumerate(headers, 1):
            cell = self.sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        self.workbook.save(self.filename)
    
    def create_widgets(self):
        # Toolbar
        toolbar = tk.Frame(self.root)
        toolbar.pack(fill=tk.X, padx=5, pady=5)
        
        tk.Button(toolbar, text="Add Entry", command=self.add_entry, 
                 bg="#007bff", fg="white").pack(side=tk.LEFT, padx=2)
        tk.Button(toolbar, text="Delete Selected", command=self.delete_entry,
                 bg="#dc3545", fg="white").pack(side=tk.LEFT, padx=2)
        tk.Button(toolbar, text="Refresh", command=self.load_data,
                 bg="#28a745", fg="white").pack(side=tk.LEFT, padx=2)
        
        # Search
        tk.Label(toolbar, text="Search:").pack(side=tk.LEFT, padx=(10,2))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.search_data())
        tk.Entry(toolbar, textvariable=self.search_var, width=30).pack(side=tk.LEFT, padx=2)
        
        # Treeview (table)
        self.tree = ttk.Treeview(self.root, columns=('date', 'main', 'main_status', 
                                                      'canada', 'canada_status', 
                                                      'youtube', 'youtube_status'), 
                                 show='headings')
        
        # Define columns
        self.tree.heading('date', text='Date')
        self.tree.heading('main', text='Main Page')
        self.tree.heading('main_status', text='Status')
        self.tree.heading('canada', text='Canada LC')
        self.tree.heading('canada_status', text='Status')
        self.tree.heading('youtube', text='YouTube')
        self.tree.heading('youtube_status', text='Status')
        
        # Set column widths
        for col in self.tree['columns']:
            self.tree.column(col, width=150)
        
        self.tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(self.root, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=scrollbar.set)
    
    def load_data(self):
        # Clear existing data
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Load from Excel
        for row in range(2, self.sheet.max_row + 1):
            values = [
                self.sheet.cell(row=row, column=1).value or '',
                self.sheet.cell(row=row, column=2).value or '',
                self.sheet.cell(row=row, column=3).value or '',
                self.sheet.cell(row=row, column=4).value or '',
                self.sheet.cell(row=row, column=5).value or '',
                self.sheet.cell(row=row, column=6).value or '',
                self.sheet.cell(row=row, column=7).value or ''
            ]
            self.tree.insert('', tk.END, values=values, iid=row)
    
    def add_entry(self):
        # Create add dialog
        dialog = tk.Toplevel(self.root)
        dialog.title("Add Entry")
        dialog.geometry("500x500")
        
        entries = {}
        fields = ['Date (DD-Mon):', 'Main Page:', 'Main Status:', 
                 'Canada LC:', 'Canada Status:', 'YouTube:', 'YouTube Status:']
        
        for i, field in enumerate(fields):
            tk.Label(dialog, text=field).pack(pady=5)
            entry = tk.Entry(dialog, width=40)
            entry.pack()
            entries[field] = entry
        
        def save():
            # Add to Excel
            next_row = self.sheet.max_row + 1
            self.sheet.cell(row=next_row, column=1, value=entries[fields[0]].get())
            self.sheet.cell(row=next_row, column=2, value=entries[fields[1]].get())
            self.sheet.cell(row=next_row, column=3, value=entries[fields[2]].get())
            self.sheet.cell(row=next_row, column=4, value=entries[fields[3]].get())
            self.sheet.cell(row=next_row, column=5, value=entries[fields[4]].get())
            self.sheet.cell(row=next_row, column=6, value=entries[fields[5]].get())
            self.sheet.cell(row=next_row, column=7, value=entries[fields[6]].get())
            self.workbook.save(self.filename)
            self.load_data()
            dialog.destroy()
        
        tk.Button(dialog, text="Save", command=save, bg="#28a745", fg="white").pack(pady=10)
    
    def delete_entry(self):
        selected = self.tree.selection()
        if selected:
            if messagebox.askyesno("Confirm", "Delete selected entry?"):
                row = int(selected[0])
                self.sheet.delete_rows(row)
                self.workbook.save(self.filename)
                self.load_data()
    
    def search_data(self):
        keyword = self.search_var.get().lower()
        if not keyword:
            self.load_data()
            return
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for row in range(2, self.sheet.max_row + 1):
            match = False
            for col in range(1, 8):
                value = str(self.sheet.cell(row=row, column=col).value or '').lower()
                if keyword in value:
                    match = True
                    break
            
            if match:
                values = [
                    self.sheet.cell(row=row, column=c).value or '' 
                    for c in range(1, 8)
                ]
                self.tree.insert('', tk.END, values=values)

if __name__ == '__main__':
    root = tk.Tk()
    app = ContentCalendarApp(root)
    root.mainloop()